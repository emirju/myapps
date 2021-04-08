
'''
Install openpyxl using pip
$ pip install openpyxl
'''

from openpyxl import load_workbook #(this will allow us to open an excel file that has been already CREATED )
from openpyxl import Workbook #(this one will allow us to create workbook )


# first thing to do to OPEN PYXL is to always create workbook instance
wb = Workbook()

#We can LOAD something that already exist
wb = load_workbook('excel_spreadsheet.xlsx') ##File name excel_spreadsheet.xlsx

work_sheet = wb.active ## an active sheet in excel file ('excel_spreadsheet.xlsx')
