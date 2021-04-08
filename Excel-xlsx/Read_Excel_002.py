from openpyxl import load_workbook

wb = load_workbook('python.xlsx')

ws = wb['Data']

row = ws.max_row
column = ws.max_column
print(f' There are {row} rows, and {column} columns')

for r in range (1, row):
    for c in range(1, column):
        print(ws.cell(r,c).value)

