
from openpyxl import load_workbook

#Creating WB instance and loading excel sheet, there are 2 ways of doing it
# wb = load_workbook('python.xlsx')
wb = load_workbook('C:\\Users\\ejukovic\\PycharmProjects\\Excel-xlsx\\python.xlsx')
print(type(wb)) ## Result :  separate CLASS which is WORKBOOK

# Print ACTIVE SHEET in file ('python.xlsx') ; result will be CURRENT ACTIVE SHEET
ws = wb.active ## ws (working sheet)
print('\n', ws)
print(wb.active.title)

# Sheet names that we have in excel
sheets = wb.sheetnames
print(sheets)

# Option 1 Reading from active sheet
print(ws.cell(2,3))  ##ws is from active sheet
print(ws.cell(2,3).value)

# Option 2 Reading from mentioned sheet
sheet_2 = wb['Programing_Language'] ### Name of second sheet in excel file
print(sheet_2.cell(2,1).value)

# Option 3 Reading
print(wb['Data']['A2'].value)

# Option 4 Reading
out = ws.cell(row=2, column=2).value
print(out)
















